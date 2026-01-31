"""
=============================================================================
PAPER CHECKING SERVICE
=============================================================================

This service grades student papers against an answer key using AI.

WHAT THIS SERVICE DOES:
1. Teacher uploads an answer key (PDF with correct answers)
2. Teacher provides student papers (PDFs - uploaded or from Google Drive)
3. For each student paper:
   - Extract student name and roll number (using AI vision)
   - Extract student answers (using OCR)
   - Compare with answer key SEMANTICALLY (meaning-based, not exact words)
   - Give marks for each answer
4. Generate an Excel sheet with all results

KEY FEATURES:
- Semantic grading: "H2O" and "water" both get marks
- Partial marks: Partially correct answers get partial marks
- Extracts handwritten names and roll numbers from papers
- Generates Excel with format: Name | Roll No | A1 | A2 | ... | Total

REUSES EXISTING CODE:
- OCRService: For extracting text from PDFs (we don't duplicate code!)
- llm_models: For calling GPT AI
- This follows DRY principle (Don't Repeat Yourself)
=============================================================================
"""

# =============================================================================
# IMPORTS - Libraries we need
# =============================================================================

import os           # For file operations (paths, delete temp files)
import json         # For converting Python objects to JSON and back
import tempfile     # For creating temporary files
import io           # For handling file-like objects in memory
from typing import List, Dict, Optional  # For type hints (helps IDE and readability)
from concurrent.futures import ThreadPoolExecutor, as_completed  # For parallel processing
from fastapi import UploadFile  # FastAPI's file upload type

# Our own modules (from this project)
from services.ocr_service import OCRService  # REUSE: For extracting text from PDFs
from llm_models.llm_models import llm, call_vision_api  # REUSE: For calling GPT AI
from prompts.checking_papers_prompts import CheckingPapersPrompts  # Our prompts
from utils.logger import log_step, log_success, log_error, log_debug  # For logging

# =============================================================================
# EXCEL LIBRARY (Optional - for generating Excel files)
# =============================================================================
# We try to import openpyxl. If it's not installed, Excel export won't work
# but the rest of the service will still work fine.

try:
    from openpyxl import Workbook  # Creates Excel workbooks
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Styling
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log_error("openpyxl not installed", "Excel export will not work. Run: pip install openpyxl")


# =============================================================================
# MAIN SERVICE CLASS
# =============================================================================

class CheckingPapersService:
    """
    Main service for checking/grading student papers.
    
    USAGE FLOW:
    1. Create instance: checker = CheckingPapersService()
    2. Call: result = await checker.check_papers_from_uploads(answer_key, student_papers)
    3. Download Excel: excel_bytes = checker.generate_results_excel(result)
    """
    
    def __init__(self):
        """
        Initialize the service.
        
        We create instances of other services we need:
        - ocr_service: For extracting text from PDFs (reusing existing code!)
        - llm: The AI model for grading
        """
        self.ocr_service = OCRService()  # Reuse OCR service (no code duplication!)
        self.llm = llm  # The GPT language model
    
    # =========================================================================
    # HELPER: JSON PARSING
    # =========================================================================
    
    def _parse_json(self, response: str) -> Dict:
        """
        Parse JSON from AI response.
        
        WHY WE NEED THIS:
        - AI sometimes returns JSON wrapped in ```json ... ``` markdown
        - We need to clean that up before parsing
        
        EXAMPLE:
        Input:  "```json\n{\"name\": \"Ali\"}\n```"
        Output: {"name": "Ali"}
        
        Args:
            response: Raw text response from AI
            
        Returns:
            Parsed dictionary (or error dict if parsing fails)
        """
        # Remove leading/trailing whitespace
        response = response.strip()
        
        # Remove markdown code block markers if present
        # Sometimes AI returns: ```json\n{...}\n```
        if response.startswith("```json"):
            response = response[7:]  # Remove "```json"
        elif response.startswith("```"):
            response = response[3:]  # Remove "```"
        if response.endswith("```"):
            response = response[:-3]  # Remove trailing "```"
        
        # Try to parse as JSON
        try:
            return json.loads(response.strip())
        except json.JSONDecodeError as e:
            # If parsing fails, log error and return error dict
            log_error("JSON parse failed", e)
            return {"error": str(e), "raw": response}
    
    # =========================================================================
    # STEP 1: PROCESS ANSWER KEY
    # =========================================================================
    
    async def process_answer_key(self, file: UploadFile) -> Dict:
        """
        Process the answer key file (uploaded by teacher).
        
        WHAT THIS DOES:
        1. Extract all text from the answer key PDF using OCR
        2. Send text to AI to understand the structure
        3. AI identifies questions, correct answers, marks per question
        4. AI determines if it's a quiz or assignment
        
        Args:
            file: The uploaded answer key file (PDF/DOCX)
            
        Returns:
            Dictionary with:
            - success: True/False
            - raw_text: The extracted text
            - parsed_answer_key: Structured questions with correct answers
            - assessment_type: "quiz" or "assignment"
            - total_questions: Count of questions
            - total_marks: Sum of all marks
        """
        log_step("Processing Answer Key", file.filename)
        
        # STEP 1: Use OCR service to extract text from the PDF
        # (We REUSE the existing OCR service - no duplicate code!)
        result = await self.ocr_service.process_uploaded_file(file)
        raw_text = result.get('raw_text', '')
        
        # Check if we got any text
        if not raw_text:
            raise ValueError("Could not extract text from answer key")
        
        # STEP 2: Send text to AI to parse and structure the answer key
        # The AI will identify each question, its correct answer, and marks
        prompt = CheckingPapersPrompts.PARSE_ANSWER_KEY.format(content=raw_text)
        
        # Create the message for the AI
        messages = [
            {"role": "system", "content": CheckingPapersPrompts.SYSTEM},  # AI's role
            {"role": "user", "content": prompt}  # Our request
        ]
        
        # Call the AI
        response = self.llm.invoke(messages)
        parsed_key = self._parse_json(response.content)
        
        log_success(f"Parsed answer key: {len(parsed_key.get('questions', []))} questions")
        
        # Return structured result
        return {
            "success": True,
            "raw_text": raw_text,
            "parsed_answer_key": parsed_key,
            "assessment_type": parsed_key.get('assessment_type', 'unknown'),
            "total_questions": len(parsed_key.get('questions', [])),
            "total_marks": parsed_key.get('total_marks', 0)
        }
    
    # =========================================================================
    # STEP 2: EXTRACT STUDENT INFO (Name, Roll Number)
    # =========================================================================
    
    def _extract_student_info_from_image(self, base64_image: str) -> Dict:
        """
        Extract student name and roll number from the first page image.
        
        WHAT THIS DOES:
        - Uses GPT Vision to "see" the first page of student's paper
        - Finds student name (even if handwritten)
        - Finds roll number / student ID
        
        WHY VISION:
        - Students often write name/roll number by hand
        - Regular OCR might miss it, but Vision AI can understand it
        
        Args:
            base64_image: The first page image encoded as base64 string
            
        Returns:
            Dictionary with student_name, roll_number, confidence
        """
        try:
            # Call GPT Vision API with the image
            response = call_vision_api(
                base64_image,  # The image
                CheckingPapersPrompts.EXTRACT_STUDENT_INFO,  # What to extract
                CheckingPapersPrompts.SYSTEM  # AI's role
            )
            return self._parse_json(response)
        except Exception as e:
            log_error("Failed to extract student info", e)
            # Return Unknown if extraction fails
            return {"student_name": "Unknown", "roll_number": "Unknown", "error": str(e)}
    
    # =========================================================================
    # STEP 3: GRADE ANSWERS (The core grading logic)
    # =========================================================================
    
    def _grade_answers(self, answer_key: Dict, student_answers: List[Dict], assessment_type: str) -> Dict:
        """
        Grade student answers against the answer key.
        
        THIS IS THE CORE GRADING FUNCTION!
        
        HOW IT WORKS:
        1. Takes the correct answers (answer_key)
        2. Takes what student wrote (student_answers)
        3. Sends both to AI for SEMANTIC comparison
        4. AI compares MEANING, not exact words
        5. AI gives marks based on understanding
        
        SEMANTIC GRADING EXAMPLES:
        - "H2O" vs "water" → Both correct (same meaning)
        - "CPU processes data" vs "The central processing unit handles data" → Both correct
        - Partial understanding → Partial marks
        
        Args:
            answer_key: The parsed answer key with correct answers
            student_answers: List of student's answers extracted by OCR
            assessment_type: "quiz" or "assignment"
            
        Returns:
            Dictionary with evaluations (marks for each answer) and totals
        """
        log_step("Grading Answers", f"Type: {assessment_type}")
        
        # Convert to JSON text for the AI prompt
        answer_key_text = json.dumps(answer_key.get('questions', []), indent=2)
        student_answers_text = json.dumps(student_answers, indent=2)
        
        # Choose the right prompt based on assessment type
        if assessment_type == "quiz":
            # Quiz: MCQ, True/False, Fill in blanks (usually full marks or zero)
            prompt = CheckingPapersPrompts.GRADE_QUIZ_ANSWERS.format(
                answer_key=answer_key_text,
                student_answers=student_answers_text
            )
        else:
            # Assignment: Long answers (can have partial marks)
            prompt = CheckingPapersPrompts.GRADE_ASSIGNMENT_ANSWERS.format(
                answer_key=answer_key_text,
                student_answers=student_answers_text,
                total_questions=len(answer_key.get('questions', []))
            )
        
        # Send to AI for grading
        messages = [
            {"role": "system", "content": CheckingPapersPrompts.SYSTEM},
            {"role": "user", "content": prompt}
        ]
        
        response = self.llm.invoke(messages)
        grading_result = self._parse_json(response.content)
        
        log_success(f"Graded: {grading_result.get('total_obtained', 0)}/{grading_result.get('total_max', 0)}")
        
        return grading_result
    
    # =========================================================================
    # STEP 4: PROCESS SINGLE STUDENT
    # =========================================================================
    
    def _process_single_student(self, file_path: str, filename: str, answer_key: Dict, assessment_type: str) -> Dict:
        """
        Process and grade a single student's paper.
        
        THIS IS THE MAIN FUNCTION FOR EACH STUDENT!
        
        STEPS:
        1. Convert first page to image
        2. Extract student name & roll number (using Vision AI)
        3. Extract all answers from the paper (using OCR)
        4. Grade the answers against answer key
        5. Return results
        
        Args:
            file_path: Path to the student's PDF file
            filename: Original filename (for display)
            answer_key: The parsed answer key
            assessment_type: "quiz" or "assignment"
            
        Returns:
            Dictionary with student info, grades, and totals
        """
        log_step("Processing Student Paper", filename)
        
        # Get file extension (.pdf, .jpg, etc.)
        suffix = os.path.splitext(file_path)[1].lower()
        
        try:
            # Initialize student info with defaults
            student_info = {"student_name": "Unknown", "roll_number": "Unknown"}
            
            # STEP 1 & 2: Extract student info from first page
            if suffix == '.pdf':
                # Convert PDF pages to images (reusing OCR service!)
                images = self.ocr_service._pdf_to_images(file_path)
                
                if images:
                    # Get first page as base64 image
                    first_page_b64 = self.ocr_service._pil_to_base64(images[0])
                    # Extract name and roll number using Vision AI
                    student_info = self._extract_student_info_from_image(first_page_b64)
            
            # STEP 3: Extract all answers from the paper (reusing OCR service!)
            file_result = self.ocr_service._process_file(file_path, suffix)
            student_answers = file_result.get('answers', [])
            
            # STEP 4: Grade the answers
            grading_result = self._grade_answers(answer_key, student_answers, assessment_type)
            
            # Return complete result for this student
            return {
                "filename": filename,
                "student_name": student_info.get('student_name', 'Unknown'),
                "roll_number": student_info.get('roll_number', 'Unknown'),
                "answers_extracted": len(student_answers),
                "grading": grading_result,
                "total_obtained": grading_result.get('total_obtained', 0),
                "total_max": grading_result.get('total_max', 0),
                "success": True
            }
            
        except Exception as e:
            # If anything fails, return error result
            log_error(f"Failed to process {filename}", e)
            return {
                "filename": filename,
                "student_name": "Unknown",
                "roll_number": "Unknown",
                "error": str(e),
                "success": False
            }
    
    # =========================================================================
    # PUBLIC METHOD 1: CHECK PAPERS FROM FILE UPLOADS
    # =========================================================================
    
    async def check_papers_from_uploads(
        self, 
        answer_key_file: UploadFile, 
        student_files: List[UploadFile]
    ) -> Dict:
        """
        Check papers when teacher uploads files directly.
        
        THIS IS THE MAIN FUNCTION CALLED BY THE API (upload method)!
        
        FLOW:
        1. Process the answer key → Get correct answers
        2. For each student file:
           a. Save to temp file
           b. Process and grade
           c. Delete temp file
        3. Compile all results
        
        Args:
            answer_key_file: The answer key PDF uploaded by teacher
            student_files: List of student paper PDFs
            
        Returns:
            Complete grading results for all students
        """
        log_step("Checking Papers (Uploads)", f"Students: {len(student_files)}")
        
        # STEP 1: Process the answer key first
        answer_key_result = await self.process_answer_key(answer_key_file)
        
        # Check if answer key processing was successful
        if not answer_key_result.get('success'):
            return {"success": False, "error": "Failed to process answer key"}
        
        # Extract what we need from answer key
        answer_key = answer_key_result['parsed_answer_key']
        assessment_type = answer_key_result['assessment_type']
        
        # STEP 2: Process each student file
        results = []
        for file in student_files:
            # Get file extension
            suffix = os.path.splitext(file.filename)[1].lower()
            
            # Save uploaded file to a temporary file on disk
            # (We need a file path to process it)
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                content = await file.read()  # Read file content
                tmp.write(content)           # Write to temp file
                tmp_path = tmp.name          # Get the path
            
            try:
                # Process and grade this student's paper
                result = self._process_single_student(
                    tmp_path, file.filename, answer_key, assessment_type
                )
                results.append(result)
            finally:
                # ALWAYS delete the temp file (even if error occurred)
                os.unlink(tmp_path)
        
        # STEP 3: Compile and return all results
        return self._compile_results(results, answer_key_result, assessment_type)
    
    # =========================================================================
    # PUBLIC METHOD 2: CHECK PAPERS FROM GOOGLE DRIVE
    # =========================================================================
    
    def check_papers_from_drive(
        self, 
        answer_key_text: str,
        drive_url: str
    ) -> Dict:
        """
        Check papers when student papers are in Google Drive.
        
        THIS IS THE MAIN FUNCTION CALLED BY THE API (Google Drive method)!
        
        FLOW:
        1. Parse the answer key text → Get correct answers
        2. Download all files from Google Drive (reusing OCR service!)
        3. Process each file in parallel (faster!)
        4. Compile all results
        
        Args:
            answer_key_text: Raw text from answer key (already extracted)
            drive_url: Google Drive folder URL with student papers
            
        Returns:
            Complete grading results for all students
        """
        log_step("Checking Papers (Google Drive)", drive_url[:50])
        
        # STEP 1: Parse the answer key text
        prompt = CheckingPapersPrompts.PARSE_ANSWER_KEY.format(content=answer_key_text)
        messages = [
            {"role": "system", "content": CheckingPapersPrompts.SYSTEM},
            {"role": "user", "content": prompt}
        ]
        response = self.llm.invoke(messages)
        answer_key = self._parse_json(response.content)
        assessment_type = answer_key.get('assessment_type', 'assignment')
        
        # Store answer key info for later
        answer_key_result = {
            "parsed_answer_key": answer_key,
            "assessment_type": assessment_type,
            "total_questions": len(answer_key.get('questions', [])),
            "total_marks": answer_key.get('total_marks', 0)
        }
        
        # STEP 2: Download files from Google Drive
        # (Reusing the function from OCR service - no code duplication!)
        file_paths = self.ocr_service._download_from_google_drive(drive_url)
        
        # Check if any files were found
        if not file_paths:
            return {"success": False, "error": "No files found in Google Drive link"}
        
        # STEP 3: Process files in PARALLEL for speed
        # (Processing 3 files at a time is faster than one-by-one)
        results = []
        
        def process_one(file_path):
            """Helper function to process one file."""
            filename = os.path.basename(file_path)  # Get just the filename
            return self._process_single_student(file_path, filename, answer_key, assessment_type)
        
        # Use ThreadPoolExecutor for parallel processing
        # max_workers=3 means process 3 files at the same time
        with ThreadPoolExecutor(max_workers=3) as executor:
            results = list(executor.map(process_one, file_paths))
        
        # STEP 4: Compile and return all results
        return self._compile_results(results, answer_key_result, assessment_type)
    
    # =========================================================================
    # HELPER: COMPILE ALL RESULTS
    # =========================================================================
    
    def _compile_results(self, results: List[Dict], answer_key_result: Dict, assessment_type: str) -> Dict:
        """
        Compile all grading results into a single response.
        
        This organizes all the data nicely for the API response.
        
        Args:
            results: List of individual student results
            answer_key_result: Info about the answer key
            assessment_type: "quiz" or "assignment"
            
        Returns:
            Organized dictionary with all results and summary
        """
        # Count successes and failures
        successful = [r for r in results if r.get('success')]
        failed = [r for r in results if not r.get('success')]
        
        return {
            "success": True,
            "assessment_type": assessment_type,
            "total_students": len(results),
            "successful": len(successful),
            "failed": len(failed),
            "answer_key_info": {
                "total_questions": answer_key_result.get('total_questions', 0),
                "total_marks": answer_key_result.get('total_marks', 0)
            },
            "results": results  # All individual student results
        }
    
    # =========================================================================
    # EXCEL GENERATION
    # =========================================================================
    
    def generate_results_excel(self, checking_results: Dict) -> bytes:
        """
        Generate Excel spreadsheet with grading results.
        
        EXCEL FORMAT FOR ASSIGNMENT:
        | S.No | Name | Roll Number | A1 | A2 | A3 | ... | Total |
        |------|------|-------------|----|----|----| ... |-------|
        | 1    | Ali  | 2021-CS-101 | 8/10 | 9/10 | ... | 42/50 |
        
        EXCEL FORMAT FOR QUIZ:
        | S.No | Name | Roll Number | Obtained / Total |
        |------|------|-------------|-----------------|
        | 1    | Ali  | 2021-CS-101 | 8 / 10          |
        
        Args:
            checking_results: Full results from check_papers_* methods
            
        Returns:
            Excel file as bytes (ready to send as download)
        """
        # Check if Excel library is installed
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
        
        log_step("Generating Excel", f"Students: {checking_results.get('total_students', 0)}")
        
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active  # Get the active sheet
        ws.title = "Grading Results"
        
        # =====================================================================
        # DEFINE STYLES (to make Excel look nice)
        # =====================================================================
        
        # Header row style: Bold, white text on blue background
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Center alignment for cells
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Thin border around cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # =====================================================================
        # PREPARE DATA
        # =====================================================================
        
        assessment_type = checking_results.get('assessment_type', 'assignment')
        results = checking_results.get('results', [])
        
        # Handle empty results
        if not results:
            ws['A1'] = "No results to display"
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        
        # Find number of questions from first successful result
        first_successful = next((r for r in results if r.get('success')), None)
        num_questions = 0
        if first_successful:
            evaluations = first_successful.get('grading', {}).get('evaluations', [])
            num_questions = len(evaluations)
        
        # =====================================================================
        # CREATE HEADERS
        # =====================================================================
        
        # Start with basic columns
        headers = ["S.No", "Name", "Roll Number"]
        
        if assessment_type == "quiz":
            # For quiz: Just one column with total marks
            headers.append("Obtained / Total")
        else:
            # For assignment: One column per answer (A1, A2, A3, ...)
            for i in range(1, num_questions + 1):
                headers.append(f"A{i}")
            headers.append("Total Obtained / Total")
        
        # Write headers to Excel (row 1)
        for col, header in enumerate(headers, 1):  # enumerate starts from 1
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # =====================================================================
        # WRITE DATA ROWS (one row per student)
        # =====================================================================
        
        for row_idx, result in enumerate(results, 2):  # Start from row 2 (row 1 is headers)
            
            # Column 1: Serial number
            ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border
            
            # Column 2: Student name
            ws.cell(row=row_idx, column=2, value=result.get('student_name', 'Unknown')).border = thin_border
            
            # Column 3: Roll number
            ws.cell(row=row_idx, column=3, value=result.get('roll_number', 'Unknown')).border = thin_border
            
            # Handle failed processing
            if not result.get('success'):
                # Show error message
                ws.cell(row=row_idx, column=4, value=f"Error: {result.get('error', 'Unknown')}").border = thin_border
                continue  # Skip to next student
            
            # Get grading data
            grading = result.get('grading', {})
            evaluations = grading.get('evaluations', [])
            total_obtained = grading.get('total_obtained', 0)
            total_max = grading.get('total_max', 0)
            
            if assessment_type == "quiz":
                # QUIZ: Just show total marks in one column
                cell = ws.cell(row=row_idx, column=4, value=f"{total_obtained} / {total_max}")
                cell.alignment = center_align
                cell.border = thin_border
            else:
                # ASSIGNMENT: Show marks for each answer
                for i, evaluation in enumerate(evaluations):
                    col_idx = 4 + i  # Start from column 4
                    obtained = evaluation.get('obtained_marks', 0)
                    max_marks = evaluation.get('max_marks', 0)
                    cell = ws.cell(row=row_idx, column=col_idx, value=f"{obtained}/{max_marks}")
                    cell.alignment = center_align
                    cell.border = thin_border
                
                # Last column: Total marks
                total_col = 4 + num_questions
                cell = ws.cell(row=row_idx, column=total_col, value=f"{total_obtained} / {total_max}")
                cell.alignment = center_align
                cell.border = thin_border
        
        # =====================================================================
        # AUTO-ADJUST COLUMN WIDTHS (so text fits nicely)
        # =====================================================================
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column letter (A, B, C, ...)
            
            # Find the longest text in this column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (with some padding, max 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # =====================================================================
        # SAVE TO BYTES AND RETURN
        # =====================================================================
        
        # Save workbook to memory (not to file)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)  # Go back to start of buffer
        
        log_success("Excel generated successfully")
        return buffer.getvalue()  # Return as bytes
